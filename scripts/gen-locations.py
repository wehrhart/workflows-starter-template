#!/usr/bin/env python3
"""Regenerate worker/data/locations.ts from a master template's "Locations" tab.

That tab is the source of truth for the column-A drop-down and carries each
facility's street address, which the resolver needs to disambiguate multiple
locations that share one zip code.

Usage:
    # Drop the master Bill-Only template next to this script (it is gitignored),
    # then run from the repo root:
    python3 scripts/gen-locations.py [path/to/template.xlsm]

Requires: pip install openpyxl
"""
import json, pathlib, re, sys
from collections import defaultdict
import openpyxl

root = pathlib.Path(__file__).resolve().parent.parent
src = pathlib.Path(sys.argv[1]) if len(sys.argv) > 1 else root / "scripts/bill-only-template.xlsm"
if not src.exists():
    sys.exit(f"template not found: {src}\nDrop the master .xlsm at that path or pass one as an argument.")

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb["Locations"]  # A=id B=name C=addr1 D=addr2 E=city F=state G=postal

def norm_addr(a: str) -> str:
    a = str(a or "").lower().strip()
    a = re.sub(r"[.,]", " ", a)
    for long, short in [("road", "rd"), ("avenue", "ave"), ("street", "st"),
                        ("drive", "dr"), ("boulevard", "blvd"), ("lane", "ln"),
                        ("court", "ct"), ("place", "pl"), ("parkway", "pkwy")]:
        a = re.sub(rf"\b{long}\b", short, a)
    return re.sub(r"\s+", " ", a).strip()

names, addrs, zips, zc = {}, {}, {}, defaultdict(list)
for r in range(2, ws.max_row + 1):
    lid = ws.cell(r, 1).value
    if lid is None:
        continue
    lid = str(lid).strip()
    names[lid] = str(ws.cell(r, 2).value or "").strip()
    addrs[lid] = norm_addr(ws.cell(r, 3).value)
    z = str(ws.cell(r, 7).value or "").strip()[:5]
    zips[lid] = z
    if z:
        zc[z].append(lid)

order = sorted(names)
names = {k: names[k] for k in order}
addrs = {k: addrs[k] for k in order}
zips = {k: zips[k] for k in order}
zc = {z: sorted(set(v)) for z, v in sorted(zc.items())}
js = lambda o: json.dumps(o, indent=2, ensure_ascii=False)
out = (
    '// AUTO-GENERATED from the master template\'s "Locations" tab. Do not edit by hand;\n'
    "// regenerate with scripts/gen-locations.py. Powers column A (Surgery Location).\n\n"
    "/** Location ID -> facility name. */\n"
    "export const LOCATION_NAMES: Record<string, string> = " + js(names) + ";\n\n"
    "/** Location ID -> normalized street address (for disambiguating shared zips). */\n"
    "export const LOCATION_ADDR: Record<string, string> = " + js(addrs) + ";\n\n"
    "/** Location ID -> 5-digit zip. */\n"
    "export const LOCATION_ZIP: Record<string, string> = " + js(zips) + ";\n\n"
    "/** Zip -> candidate Location IDs at that zip. */\n"
    "export const ZIP_CANDIDATES: Record<string, string[]> = " + js(zc) + ";\n"
)
(root / "worker/data/locations.ts").write_text(out)
print(f"wrote worker/data/locations.ts  locations={len(names)} zips={len(zc)}")
